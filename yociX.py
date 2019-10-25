import openpyxl
import re
import traceback


changeCells = 0

# replace the special content
"""
file: file path : str
mode: type of the operatoration : int
text: the string need to be replaceed : int or str
replaceText: replacement Text : int or str
"""


def changeData(file, mode, text, replaceText):
    # load the file(*.xlsx)
    wb = openpyxl.load_workbook(file)
    # ! deal with one sheet
    ws = wb.worksheets[0]
    global changeCells
    # get rows and columns of file
    rows = ws.max_row
    cols = ws.max_column
    changeFlag = False
    try:
        for row in range(1, rows+1):
            for col in range(1, cols+1):
                content = ws.cell(row=row, column=col).value
                if(content != None):
                    # mode1: fullmatch replacement
                    if(mode == 1):
                        if(content == text):
                            ws.cell(row=row, column=col).value = replaceText
                            changeFlag = True
                            changeCells += 1
                    # mode2: partial replacement
                    elif(mode == 2):
                        if(type(content) == str):
                            ws.cell(row=row, column=col).value = content.replace(
                                text, replaceText, 1)
                            changeFlag = True
                            changeCells += 1
                    # mode3: partialmatch and filling
                    elif(mode == 3):
                        if(type(content) == str):
                            ws.cell(row=row, column=col).value = content.replace(
                                text, text+replaceText, 1)
                            changeFlag = True
                            changeCells += 1
                    else:
                        return 0
        # status_1: modified success
        if(changeFlag):
            wb.save(file)
            return changeCells
        # status_2: no modified
        else:
            return changeCells
    # status_3: exception
    except Exception as e:
        print(traceback.format_exc())


# read the content of file
"""
file: file path : str
"""


def rdxl(file):
    # load the file(*.xlsx)
    wb = openpyxl.load_workbook(file)
    # ! deal with one sheet
    ws = wb.worksheets[0]
    global changeCells
    # get rows and columns of file
    rows = ws.max_row
    cols = ws.max_column
    changeFlag = False
    cells = 0
    for row in range(1, rows+1):
        for col in range(1, cols+1):
            content = ws.cell(row=row, column=col).value
            print(content)
            cells += 1
    print('cells', cells)


# if __name__ == "__main__":
#     res = changeData('D:\\001.xlsx', 1, 7777, 'bug制造者')
#     if(res != None):
#         print('已修改 ', res, ' 处')
#     # else:
#     #     print('操作失败:\n'+res)
#     rdxl('D:\\001.xlsx')

# 10.17初步实现了文件的全字符匹配替换；部分匹配替换；部分匹配填充；
